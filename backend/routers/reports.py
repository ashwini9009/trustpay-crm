from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from database import get_db
from models.models import Partner, Target
from utils.auth import require_admin
import io
import csv
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
from routers import reports 

router = APIRouter(prefix="/api/reports", tags=["Reports"])

def get_partners_data(db: Session):
    partners = db.query(Partner).all()
    data = []
    for p in partners:
        targets = p.targets
        total = len(targets)
        completed = sum(1 for t in targets if t.is_completed)
        data.append({
            "name": p.user.name,
            "email": p.user.email,
            "company": p.company_name or "N/A",
            "city": p.city or "N/A",
            "phone": p.phone or "N/A",
            "business_type": p.business_type or "N/A",
            "total_targets": total,
            "completed_targets": completed,
            "completion_rate": f"{(completed/total*100):.1f}%" if total > 0 else "0%",
            "is_active": "Active" if p.is_active else "Inactive",
            "joining_date": p.joining_date.strftime("%d %b %Y") if p.joining_date else "N/A",
        })
    return data

# ✅ PDF Report
@router.get("/pdf")
def download_pdf(db: Session = Depends(get_db), admin=Depends(require_admin)):
    partners = get_partners_data(db)
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=30, rightMargin=30)
    styles = getSampleStyleSheet()
    elements = []

    # Title
    title = Paragraph("<b>TrustPay Loans — Partner Report</b>", styles["Title"])
    date = Paragraph(f"Generated on: {datetime.now().strftime('%d %B %Y %I:%M %p')}", styles["Normal"])
    elements.append(title)
    elements.append(date)
    elements.append(Spacer(1, 20))

    # Table headers
    headers = ["Name", "Company", "City", "Phone", "Targets", "Completed", "Rate", "Status"]
    table_data = [headers]

    for p in partners:
        table_data.append([
            p["name"],
            p["company"],
            p["city"],
            p["phone"],
            str(p["total_targets"]),
            str(p["completed_targets"]),
            p["completion_rate"],
            p["is_active"],
        ])

    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a1a2e")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f4f7fb")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("PADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(table)

    # Summary
    elements.append(Spacer(1, 20))
    total_partners = len(partners)
    active = sum(1 for p in partners if p["is_active"] == "Active")
    total_targets = sum(p["total_targets"] for p in partners)
    completed_targets = sum(p["completed_targets"] for p in partners)
    summary = Paragraph(
        f"<b>Summary:</b> Total Partners: {total_partners} | Active: {active} | "
        f"Total Targets: {total_targets} | Completed: {completed_targets}",
        styles["Normal"]
    )
    elements.append(summary)

    doc.build(elements)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=partners_report_{datetime.now().strftime('%Y%m%d')}.pdf"}
    )

# ✅ Excel Report
@router.get("/excel")
def download_excel(db: Session = Depends(get_db), admin=Depends(require_admin)):
    partners = get_partners_data(db)
    wb = Workbook()
    ws = wb.active
    ws.title = "Partners Report"

    # Header style
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    headers = ["Name", "Email", "Company", "City", "Phone", "Business Type",
               "Total Targets", "Completed", "Completion Rate", "Status", "Joining Date"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row, p in enumerate(partners, 2):
        values = [
            p["name"], p["email"], p["company"], p["city"], p["phone"],
            p["business_type"], p["total_targets"], p["completed_targets"],
            p["completion_rate"], p["is_active"], p["joining_date"]
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(horizontal="center")
            if row % 2 == 0:
                cell.fill = PatternFill(start_color="f4f7fb", end_color="f4f7fb", fill_type="solid")

    # Auto column width
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 4

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "TrustPay Loans - Summary"
    ws2["A1"].font = Font(bold=True, size=14)
    ws2["A3"] = "Total Partners"
    ws2["B3"] = len(partners)
    ws2["A4"] = "Active Partners"
    ws2["B4"] = sum(1 for p in partners if p["is_active"] == "Active")
    ws2["A5"] = "Total Targets"
    ws2["B5"] = sum(p["total_targets"] for p in partners)
    ws2["A6"] = "Completed Targets"
    ws2["B6"] = sum(p["completed_targets"] for p in partners)
    ws2["A7"] = "Generated On"
    ws2["B7"] = datetime.now().strftime("%d %B %Y %I:%M %p")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=partners_report_{datetime.now().strftime('%Y%m%d')}.xlsx"}
    )

# ✅ CSV Report
@router.get("/csv")
def download_csv(db: Session = Depends(get_db), admin=Depends(require_admin)):
    partners = get_partners_data(db)
    buffer = io.StringIO()
    writer = csv.DictWriter(buffer, fieldnames=partners[0].keys() if partners else [])
    writer.writeheader()
    writer.writerows(partners)
    buffer.seek(0)
    return StreamingResponse(
        io.BytesIO(buffer.getvalue().encode()),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=partners_report_{datetime.now().strftime('%Y%m%d')}.csv"}
    )